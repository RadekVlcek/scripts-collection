# DISCLAIMER: This script is not functional anymore and serves only for pleasure of your eyes

# Description: Read data from XML configuration file and store it in Excel workbook

import os
from shutil import copyfile
from datetime import datetime
from bs4 import BeautifulSoup as bs
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Color, colors, PatternFill, Border, Side, Alignment, Protection, Font

log = '> '

original_file = '\\\\imaginary_path\\imaginary_path\\imaginary_path\\imaginary_path\\imaginary_file.xml'
temp_file = '\\\\imaginary_path2\\imaginary_path2\\imaginary_path2\\imaginary_path2\\imaginary_file2.xml'

# Make temp copy of config file
try:
    copyfile(original_file, temp_file)
    print(f'{log}Making temp copy of {original_file} file')
except IOError as e:
    print(e)

# Parse XML file
content = []
with open(temp_file, 'r') as file:
    content = file.readlines()
    content = "".join(content)
    bs_content = bs(content, "lxml")

route_collection = []
all_routes = bs_content.find_all('route')

# Collect data from XML
print(f'{log}Reading data')

for route in all_routes:
    r_name = route.get('name')
    r_name = r_name.split('_')
    
    try:
        r_name.remove('RT')
    except Exception as e:
        print(e)
        
    try:
        r_name.remove('SIP')
    except Exception as e:
        print(e)
        
    r_name = ' '.join(r_name)
    
    r_prio = route.find('priority')
    r_prio = int(r_prio.text) if r_prio != None else '(No priority)'
    
    if r_prio == 1:
        r_prio = f'{r_prio} (in use)'
    else:
        r_prio = f'{r_prio} (not in use)'

    r_cli = route.find('presentation').find('originating_addr')
    r_cli = r_cli.text if r_cli != None else '(No CLI)'
    
    r_trunk_ip = route.find('postfix')
    r_trunk_ip = r_trunk_ip.text if r_trunk_ip != None else '(No SIP trunk)'
    
    # Append collected information to list
    r_collection.append((r_name, r_prio, r_cli, r_trunk_ip))
    
# Convert list to tuple
r_collection = tuple(r_collection)

# Initiate Excel workbook
print(f'{log}Initiating workbook')
book = Workbook()
sheet = book.active
sheet.title = 'Config file'

# Define some sheet parameters
font = Font(size=13, color="00FFFFFF", bold=True)
alignment = Alignment(horizontal='center', vertical='center')
border = Border(left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
fill = PatternFill(fill_type='solid', start_color='0016365c', end_color='00000000')

# Set and print INTRO information
now = datetime.now().strftime("%d.%m.%Y at %H:%M CET")
sheet.cell(row=1, column=1).value = 'Overview of the currently used XXXX'
sheet.cell(row=2, column=1).value = 'This sheet gets updated automatically every 24 hours'
sheet.cell(row=3, column=1).value = f'Last updated: {now}'
sheet.cell(row=4, column=1).value = 'Each country uses two providers: XXXX and XXXXX'
sheet.cell(row=5, column=1).value = 'CTRL + F to search for a country'
sheet.cell(row=6, column=1).value = "Check 'Priority' column to see which provider is in use"
sheet.cell(row=7, column=1).value = "Check 'Current CLI' column to see CLI used for the country"

# Style INTRO information
sheet.cell(row=1, column=1).font = Font(size=16, bold=True)
sheet.cell(row=3, column=1).border = Border(bottom=Side(border_style='thin', color='FF000000'))
sheet.row_dimensions[1].height = 20
for i in range(4, 8):
    sheet.cell(row=i, column=1).font = Font(size=9)
    sheet.cell(row=i, column=1).border = Border(right=Side(border_style='thin', color='FF000000'))

# Set and print HEADERS
headers = ['SIP provider - Route/Country', 'Priority', 'Current CLI', 'SIP Trunk IP']
for i in range(4):
    sheet.cell(row=8, column=i+1).value = headers[i]
    sheet.cell(row=8, column=i+1).font = font
    sheet.cell(row=8, column=i+1).alignment = alignment
    sheet.cell(row=8, column=i+1).border = border
    sheet.cell(row=8, column=i+1).fill = fill

# Write rest of the data into the sheet
print(f'{log}Writing to workbook')
for route in r_collection:
    sheet.append(route)

# Set width for columns
sheet.column_dimensions[get_column_letter(1)].width = 50
sheet.column_dimensions[get_column_letter(2)].width = 20
sheet.column_dimensions[get_column_letter(3)].width = 40
sheet.column_dimensions[get_column_letter(4)].width = 25

# Set height for all rows
for i in range(8, len(all_routes)):
    sheet.row_dimensions[i+1].height = 25

# Set properties to all data cells
all_rows = sheet.max_row
all_columns = sheet.max_column
for row in range(9, all_rows+1):
    for column in range(1, all_columns+1):
        sheet.cell(row=row, column=column).alignment = alignment
        sheet.cell(row=row, column=column).border = border

# Where to save excel sheet
path1 = f'\\\\IMAGINARY_SERVER\\IMAGINARY_DRIVE$\\IMAGINARY_PATH\\IMAGINARY_PATH\\IMAGINARY_PATH\\IMAGINARY_FILE.xlsx'
path2 = f'\\\\IMAGINARY_SERVER2\\IMAGINARY_DRIVE$2\\IMAGINARY_PATH2\\IMAGINARY_PATH2\\IMAGINARY_PATH2\\IMAGINARY_FILE2.xlsx'

# Save excel sheet in MANAGEMENT folder
book.save(filename=path1)
print(f'{log}Saving workbook in imaginary folder')

# Save excel sheet in ICT folder
book.save(filename=path2)
print(f'{log}Saving workbook in another imaginary folder')

# Delete temp file
print(f'{log}Removing temp file')
os.remove(temp_file)

print(f'{log}Done')